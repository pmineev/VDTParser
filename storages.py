from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class LeaderboardsStorage:
    def __init__(self):
        self.nickname = None
        self.leaderboards = None
        self.deltas = None

    def save(self, path):
        workbook = Workbook()

        sheet = workbook.active

        sheet.append([
            'Дата',
            'Результат',
            'Рекорд по VDT',
            'Абс. дельта по VDT',
            'Отн. дельта по VDT',
            'Рекорд по миру',
            'Абс. дельта по миру',
            'Отн. дельта по миру'
        ])

        columns = [
            self.deltas['dates'],
            [[f for f in lb.flights if f.player == self.nickname][0].time for lb in self.leaderboards],
            [lb.flights[0].time for lb in self.leaderboards],
            self.deltas['abs_vdt'],
            self.deltas['rel_vdt'],
            [lb.world_record for lb in self.leaderboards],
            self.deltas['abs_world'],
            self.deltas['rel_world'],
        ]

        for row in zip(*columns):
            sheet.append(row)

        for col in range(1, 9):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

        workbook.save(path)
